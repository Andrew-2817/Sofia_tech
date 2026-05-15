#!/usr/bin/env python3
"""Загрузка товаров Nivona из Excel файла в БД - правильное сопоставление фото"""

import sys
import re
import shutil
from pathlib import Path
from PIL import Image
import io
from openpyxl import load_workbook

# Пути
BASE_DIR = Path('/home/raul/projects/sofa2/Sofia_tech')
EXCEL_FILE = BASE_DIR / 'Backend' / 'app' / 'zavoz' / 'exelfiles' / 'файл_товары_12.xlsx'
PHOTOS_DIR = BASE_DIR / 'Backend' / 'app' / 'zavoz' / 'photos'
UPLOADS_DIR = BASE_DIR / 'Backend' / 'static' / 'uploads' / 'products'

# Добавляем путь для импорта
sys.path.insert(0, str(BASE_DIR))

from Backend.app.database import SessionLocal
from Backend.app.models.product_nivona import NivonaProduct
from Backend.app.models.category import Category
from Backend.app.models.brand import Brand
import pandas as pd

def convert_to_jpg(image_data, quality=85):
    """Конвертирует изображение в JPG формат"""
    try:
        img = Image.open(io.BytesIO(image_data))
        if img.mode in ('RGBA', 'LA', 'P'):
            rgb_img = Image.new('RGB', img.size, (255, 255, 255))
            rgb_img.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = rgb_img
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        output = io.BytesIO()
        img.save(output, format='JPEG', quality=quality, optimize=True)
        return output.getvalue()
    except Exception as e:
        print(f"      ⚠️ Ошибка конвертации: {e}")
        return image_data

def clean_price(price_str):
    """Очистка цены от символов"""
    if pd.isna(price_str):
        return None
    price_str = str(price_str).strip()
    price_str = price_str.replace('₽', '').replace('руб', '').replace(' ', '').replace(',', '.')
    price_str = re.sub(r'[^\d.]', '', price_str)
    try:
        return float(price_str)
    except:
        return None

def clean_string(value):
    if pd.isna(value):
        return None
    return str(value).strip()

def clean_int(value):
    if pd.isna(value):
        return None
    try:
        return int(float(value))
    except:
        return None

def extract_images_from_excel():
    """Извлечение изображений - только те, что в строках с Артикулом"""
    print("=" * 70)
    print("📸 ИЗВЛЕЧЕНИЕ ФОТО NIVONA")
    print("=" * 70)

    if not EXCEL_FILE.exists():
        print(f"❌ Файл не найден: {EXCEL_FILE}")
        return {}

    # Очищаем папки перед загрузкой
    if PHOTOS_DIR.exists():
        for f in PHOTOS_DIR.glob("*.jpg"):
            f.unlink()
    if UPLOADS_DIR.exists():
        for f in UPLOADS_DIR.glob("12.*.jpg"):
            f.unlink()

    PHOTOS_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

    row_to_photo = {}
    image_counter = 0

    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active

        # Находим колонку с Артикулом (поищем в первой строке)
        sku_col = None
        for col in range(1, 20):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and 'Артикул' in str(cell_value):
                sku_col = col
                break

        if not sku_col:
            sku_col = 5  # по умолчанию колонка E

        print(f"📋 Колонка Артикул: {sku_col}")

        images = sheet._images if hasattr(sheet, '_images') else []
        print(f"🖼️ Найдено всего изображений: {len(images)}")

        # Сортируем изображения по строке
        images_by_row = {}
        for img in images:
            try:
                anchor = img.anchor
                if hasattr(anchor, '_from'):
                    row_num = anchor._from.row + 1
                    if row_num >= 2:  # только строки с данными
                        images_by_row.setdefault(row_num, []).append(img)
            except:
                pass

        # Для каждой строки, где есть Артикул, берем первое изображение
        for row_num in sorted(images_by_row.keys()):
            # Проверяем, есть ли в строке Артикул
            sku_cell = sheet.cell(row=row_num, column=sku_col).value
            if not sku_cell or sku_cell == 'Артикул':
                continue

            # Берем первое изображение в строке
            img = images_by_row[row_num][0]
            img_data = img._data()

            if len(img_data) > 100:
                converted_data = convert_to_jpg(img_data)

                image_counter += 1
                img_name = f"12.{image_counter}.jpg"
                img_path = PHOTOS_DIR / img_name

                with open(img_path, 'wb') as f:
                    f.write(converted_data)

                dest_path = UPLOADS_DIR / img_name
                shutil.copy2(img_path, dest_path)

                row_to_photo[row_num] = f"/uploads/products/{img_name}"
                print(f"  ✅ Строка {row_num}: {sku_cell} -> {img_name}")

    except Exception as e:
        print(f"❌ Ошибка при открытии файла: {e}")
        return {}

    print(f"\n📊 Итого фото товаров: {image_counter}")
    return row_to_photo

def load_products_to_db(row_to_photo):
    """Загрузка товаров Nivona в базу данных"""
    print("\n" + "=" * 70)
    print("📦 ЗАГРУЗКА ТОВАРОВ NIVONA В БАЗУ ДАННЫХ")
    print("=" * 70)

    if not EXCEL_FILE.exists():
        print(f"❌ Файл не найден: {EXCEL_FILE}")
        return

    df = pd.read_excel(EXCEL_FILE)
    print(f"📊 Найдено строк с данными: {len(df)}")
    print(f"📋 Колонки в Excel: {list(df.columns)}")

    db = SessionLocal()

    try:
        # Бренд Nivona id=12
        brand = db.query(Brand).filter(Brand.id == 12).first()
        if not brand:
            brand = Brand(id=12, name='Nivona')
            db.add(brand)
            db.commit()
            db.refresh(brand)
            print(f"✅ Создан бренд: Nivona (id: {brand.id})")
        else:
            print(f"✅ Бренд найден: Nivona (id: {brand.id})")

        new_count = 0
        photo_count = 0

        for idx, row in df.iterrows():
            excel_row_num = idx + 2

            sku = clean_string(row.get('Артикул'))
            if not sku or sku == 'nan':
                continue

            # Получаем фото по номеру строки
            photo_path = row_to_photo.get(excel_row_num)
            if photo_path:
                photo_count += 1
                print(f"  📷 {sku}: фото {photo_path.split('/')[-1]}")
            else:
                print(f"  ℹ️ {sku}: без фото")

            category_id = clean_int(row.get('category_id'))

            product_data = {
                'category_id': category_id,
                'brand_id': brand.id,
                'main_image': photo_path,
                'sku': sku,
                'model': clean_string(row.get('Модель')),
                'name': clean_string(row.get('Название')) or sku,
                'description': clean_string(row.get('Описание')),
                'price_public': clean_price(row.get('Цена')),
                'comment': clean_string(row.get('Комментарий'))
            }

            new_product = NivonaProduct(**product_data)
            db.add(new_product)
            new_count += 1

        db.commit()

        print("\n" + "=" * 70)
        print("📊 РЕЗУЛЬТАТ:")
        print(f"   ✅ ЗАГРУЖЕНО: {new_count} товаров")
        print(f"   📷 Фото добавлено: {photo_count}")
        print(f"   📦 Всего товаров в БД: {db.query(NivonaProduct).count()}")

    except Exception as e:
        print(f"❌ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    print("\n" + "=" * 70)
    print("🚀 ЗАПУСК ЗАГРУЗКИ NIVONA")
    print("=" * 70)

    row_to_photo = extract_images_from_excel()
    load_products_to_db(row_to_photo)

    print("\n" + "=" * 70)
    print("🏁 ЗАВЕРШЕНО")
    print("=" * 70)