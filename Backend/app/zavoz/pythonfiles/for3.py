#!/usr/bin/env python3
import sys
from pathlib import Path
from openpyxl import load_workbook
from PIL import Image
import io
import shutil
from collections import defaultdict

sys.path.insert(0, 'C:/vs code/Sofia_tech')

from Backend.app.database import SessionLocal, engine
from Backend.app.models.product_ilve import IlveProduct
from Backend.app.models.brand import Brand
from sqlalchemy import text
import pandas as pd
BASE_DIR = Path('/home/raul/projects/sofa2/Sofia_tech')

EXCEL_FILE = BASE_DIR / 'Backend' / 'app' / 'zavoz' / 'exelfiles' / 'файл_товары_3.xlsx'
PHOTOS_DIR = BASE_DIR / 'Backend' / 'app' / 'zavoz' / 'photos'
UPLOADS_DIR = BASE_DIR / 'Backend' / 'static' / 'uploads' / 'products'
ILVE_BRAND_ID = 14

def convert_to_jpg(img_data):
    try:
        img = Image.open(io.BytesIO(img_data))
        if img.mode != 'RGB':
            img = img.convert('RGB')
        out = io.BytesIO()
        img.save(out, format='JPEG', quality=85)
        return out.getvalue()
    except:
        return img_data

def drop_fk():
    print("🛠 Удаление foreign key constraint на category_id...")
    with engine.connect() as conn:
        try:
            conn.execute(text("ALTER TABLE products_ilve DROP CONSTRAINT IF EXISTS products_ilve_category_id_fkey"))
            conn.commit()
            print("✅ FK удалён")
        except Exception as e:
            print(f"⚠️ FK уже нет или ошибка: {e}")

def main():
    print("📸 Извлечение фото и загрузка товаров ILVE")
    PHOTOS_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    for f in PHOTOS_DIR.glob("ilve_*.jpg"): f.unlink()
    for f in UPLOADS_DIR.glob("ilve_*.jpg"): f.unlink()

    wb = load_workbook(EXCEL_FILE, data_only=False)
    sheet_names = wb.sheetnames
    print(f"Листы: {sheet_names}")

    photo_map = defaultdict(list)
    for sheet_idx, sheet_name in enumerate(sheet_names, start=1):
        ws = wb[sheet_name]
        print(f"Обработка листа '{sheet_name}'")
        if hasattr(ws, '_images') and ws._images:
            for img in ws._images:
                anchor = img.anchor
                if anchor is None: continue
                row = anchor._from.row + 1
                img_data = img._data()
                if img_data:
                    converted = convert_to_jpg(img_data)
                    img_name = f"ilve_{sheet_idx}_{row}.jpg"
                    (PHOTOS_DIR / img_name).write_bytes(converted)
                    shutil.copy(PHOTOS_DIR / img_name, UPLOADS_DIR / img_name)
                    photo_map[(sheet_name, row)].append(f"/uploads/products/{img_name}")
                    print(f"  ✅ Фото: строка {row}")
        else:
            print(f"  ⚠️ Нет изображений")

    print(f"Всего фото: {sum(len(v) for v in photo_map.values())}")

    # Отключаем FK перед вставкой
    drop_fk()

    db = SessionLocal()
    try:
        db.query(IlveProduct).delete()
        db.commit()
        print("Старые товары удалены")

        brand = db.query(Brand).filter(Brand.id == ILVE_BRAND_ID).first()
        if not brand:
            brand = Brand(id=ILVE_BRAND_ID, name='ILVE')
            db.add(brand)
            db.commit()
        print(f"Бренд: {brand.name}")

        xl = pd.ExcelFile(EXCEL_FILE)
        total = 0

        for sheet_name in xl.sheet_names:
            print(f"\n--- Лист: {sheet_name}")
            df_raw = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, header=None)
            header_row = 0
            for i in range(min(20, len(df_raw))):
                vals = [str(v).lower() for v in df_raw.iloc[i].values if pd.notna(v)]
                if any(x in vals for x in ['модель', 'наименование товара', 'ррц']):
                    header_row = i
                    break
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, header=header_row)
            print(f"  Строк данных: {len(df)}")

            for idx, row in df.iterrows():
                excel_row = idx + header_row + 2
                name = row.get('Наименование товара')
                if pd.isna(name):
                    continue

                photos = photo_map.get((sheet_name, excel_row), [])
                main_image = photos[0] if photos else None

                cat_val = row.get('ID_категории')
                cat_id = int(cat_val) if cat_val and str(cat_val).isdigit() else None

                price_val = row.get('РРЦ Руб')
                price = 0
                if pd.notna(price_val):
                    price = float(str(price_val).strip().replace('₽', '').replace(' ', '').replace(',', '.'))

                model = row.get('Модель')
                sku = str(model) if pd.notna(model) else f"ILVE-{sheet_name[:3]}-{excel_row}"

                product = IlveProduct(
                    category_id=cat_id,
                    brand_id=brand.id,
                    sku=sku,
                    model=model,
                    group=row.get('Группа товара') if pd.notna(row.get('Группа товара')) else None,
                    name=name,
                    series=row.get('Серия') if pd.notna(row.get('Серия')) else None,
                    color=row.get('Цвет прибора') if pd.notna(row.get('Цвет прибора')) else None,
                    decor_color=row.get('Цвет декора') if pd.notna(row.get('Цвет декора')) else None,
                    width=row.get('Ширина прибора') if pd.notna(row.get('Ширина прибора')) else None,
                    hob=row.get('Варочная поверхность') if pd.notna(row.get('Варочная поверхность')) else None,
                    hob_sketch=row.get('Эскиз варочной') if pd.notna(row.get('Эскиз варочной')) else None,
                    oven=row.get('Духовой шкаф') if pd.notna(row.get('Духовой шкаф')) else None,
                    price=price,
                    main_image=main_image,
                    status=row.get('Статус') if pd.notna(row.get('Статус')) else None,
                    description=row.get('Описание') if pd.notna(row.get('Описание')) else None,
                    ean=row.get('EAN') if pd.notna(row.get('EAN')) else None,
                    comment=row.get('Комментарий') if pd.notna(row.get('Комментарий')) else None
                )
                db.add(product)
                total += 1
                if total % 100 == 0:
                    db.commit()
                    print(f"  Добавлено {total} товаров")
            db.commit()
        print(f"\n✅ Добавлено {total} товаров с фото")
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    main()